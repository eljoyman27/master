# ==== SAMPLE DATA + DUMMY TEMPLATE + TEST CALLS ====
import random, string
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook

random.seed(42)

# ---------- helpers to make sample IDs ----------
def _rand_id():
    return "M" + "".join(random.choices(string.digits, k=7))

def _mk_ids(n):
    return [_rand_id() for _ in range(n)]

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
REGIONS = [f"Region {i}" for i in range(1,9)]
IDENTITY_CATS = ["gender at birth","identified as","sexual orientation","race","ethnicity"]
SUBCATS = {
    "gender at birth": ["Male","Female","Intersex"],
    "identified as": ["Man","Woman","Non-binary"],
    "sexual orientation": ["Heterosexual","Homosexual","Bisexual","Asexual"],
    "race": ["White","Black","Asian","Native","Multiracial"],
    "ethnicity": ["Hispanic","Non-Hispanic"]
}
NEEDS = ["Housing","Food","Transportation","Medicine"]

# ---------- Tab 2 sample: region × 4 HRA metrics by month ----------
def sample_tab2_rows(n_per=(2,3)):  # (regions per metric per month, members per bucket)
    rows = []
    metrics = [
        "Initial HRA Enrollees #",
        "Initial HRA Completion #",
        "Annual HRA Qualified #",
        "Annual HRA Completed #",
    ]
    r_per, members_per = n_per
    for m in MONTHS:
        for metric in metrics:
            for region in random.sample(REGIONS, k=min(r_per, len(REGIONS))):
                for _ in range(members_per):
                    rows.append({
                        "region": region,
                        "member_category": metric,   # <-- your upstream name
                        "months": m,
                        "seq_memb_id": _rand_id()
                    })
    return pd.DataFrame(rows)

# ---------- Tab 3 sample: region × identity category × subcategory by month ----------
def sample_tab3_rows(n_pairs=2):
    rows = []
    for m in MONTHS:
        for region in random.sample(REGIONS, k=3):
            for cat in IDENTITY_CATS:
                for sub in random.sample(SUBCATS[cat], k=min(n_pairs, len(SUBCATS[cat]))):
                    # 0–3 members to create some zeroes too
                    for _ in range(random.randint(0,3)):
                        rows.append({
                            "region": region,
                            "category_type": cat,      # <-- your upstream name
                            "sub_category": sub,       # <-- your upstream name
                            "months": m,
                            "seq_memb_id": _rand_id()
                        })
    return pd.DataFrame(rows)

# ---------- Tab 4 sample: age × identity category × subcategory by month ----------
def sample_tab4_rows():
    rows = []
    for m in MONTHS:
        for cat in IDENTITY_CATS:
            for sub in random.sample(SUBCATS[cat], k=min(2, len(SUBCATS[cat]))):
                for _ in range(random.randint(1,4)):
                    rows.append({
                        "age": random.choice([9, 16, 22, 28, 37, 46, 57, 60, 71, "25-34"]),  # ints or pre-bucket
                        "category_type": cat,
                        "sub_category": sub,
                        "months": m,
                        "seq_memb_id": _rand_id()
                    })
    return pd.DataFrame(rows)

# ---------- Tab 5 sample: region × need category by month ----------
def sample_tab5_rows():
    rows = []
    for m in MONTHS:
        for region in random.sample(REGIONS, k=4):
            for need in NEEDS:
                for _ in range(random.randint(0,3)):
                    rows.append({
                        "region": region,
                        "HRA NEEDS ASSESSMENT METRICS": need,  # <-- your upstream header
                        "months": m,
                        "seq_memb_id": _rand_id()
                    })
    return pd.DataFrame(rows)

# ---------- Tab 6 sample: just month + id (can reuse a union of above) ----------
def sample_tab6_rows():
    # union-like pool to simulate many IDs across months
    rows = []
    for m in MONTHS:
        for _ in range(random.randint(20,40)):
            rows.append({"months": m, "seq_memb_id": _rand_id()})
    return pd.DataFrame(rows)

# ---------- Create a dummy template with months & merged cells ----------
def create_dummy_template(path: str):
    wb = Workbook()

    # Tab 2
    ws2 = wb.active
    ws2.title = "Tab 2"
    # Month header on row 12 (Jan..Dec from col D=4)
    ws2.cell(12,1,"Region")
    ws2.cell(12,2,"Metric")
    for i,m in enumerate(MONTHS):
        ws2.cell(12,4+i, m)
    # Label block rows 13..22 (8-10 rows for testing)
    r = 13
    for reg in ["Region 1","Region 2","Region 3"]:
        for metric in [
            "Initial HRA Enrollees #",
            "Initial HRA Completion #",
            "Annual HRA Qualified #",
            "Annual HRA Completed #",
        ]:
            ws2.cell(r,1,reg)
            ws2.cell(r,2,metric)
            r += 1
        r += 1  # blank row between regions
    # Merge a couple month cells vertically to test merged-cell write
    ws2.merge_cells(start_row=14, start_column=4, end_row=15, end_column=4)  # Jan for two rows

    # Tab 3 (no header row; Jan starts at F=6)
    ws3 = wb.create_sheet("HRA Population Metrics")
    ws3.cell(89,1,"Region"); ws3.cell(89,2,"Category"); ws3.cell(89,3,"Subcategory")
    rr = 90
    for reg in ["Region 1","Region 2"]:
        ws3.cell(rr,1,reg)
        for cat in ["Gender at Birth","Identified As"]:
            ws3.cell(rr,2,cat)
            for sub in ["A","B","C"]:
                ws3.cell(rr,3,sub); rr += 1
        rr += 1  # blank
    # merge to test writer
    ws3.merge_cells(start_row=91, start_column=6, end_row=92, end_column=6)  # Jan col F merged

    # Tab 4 (header row at 30)
    ws4 = wb.create_sheet("Tab 4")
    ws4.cell(30,1,"Age Range"); ws4.cell(30,2,"Category"); ws4.cell(30,3,"Subcategory")
    for i,m in enumerate(MONTHS):
        ws4.cell(30,4+i, m)
    rr = 31
    for ar in ["0-17","18-24","25-34"]:
        ws4.cell(rr,1,ar)
        for cat in ["Gender at Birth","Ethnicity"]:
            ws4.cell(rr,2,cat)
            for sub in ["X","Y"]:
                ws4.cell(rr,3,sub); rr += 1
        rr += 1
    ws4.merge_cells(start_row=33, start_column=4, end_row=34, end_column=4)

    # Tab 5 (no header; Jan at D=4)
    ws5 = wb.create_sheet("Tab 5 - Region Needs")
    ws5.cell(49,1,"Region"); ws5.cell(49,2,"Need")
    rr = 50
    for reg in ["Region 1","Region 2","Region 3"]:
        ws5.cell(rr,1,reg)
        for need in ["Housing","Food","Transportation","Medicine"]:
            ws5.cell(rr,2,need); rr += 1
        rr += 1
    ws5.merge_cells(start_row=52, start_column=4, end_row=53, end_column=4)

    # Tab 6 (totals row test; Jan at F=6)
    ws6 = wb.create_sheet("Tab 6 - Totals")
    ws6.cell(12,6,"Jan"); ws6.cell(12,7,"Feb"); ws6.cell(12,8,"Mar")
    for i,m in enumerate(MONTHS):
        ws6.cell(12,6+i, m)
    # a merged target cell area to test writing
    ws6.merge_cells(start_row=85, start_column=6, end_row=86, end_column=6)

    wb.save(path)

# ---------- Build sample dataframes ----------
df_tab2  = sample_tab2_rows()
df_tab3  = sample_tab3_rows()
df_tab4  = sample_tab4_rows()
df_tab5  = sample_tab5_rows()
df_tab6  = sample_tab6_rows()

# ---------- Make a dummy template ----------
TEMPLATE = "dummy_template.xlsx"
OUTPUT   = "dummy_filled.xlsx"
create_dummy_template(TEMPLATE)

# ---------- NOW CALL YOUR FUNCTIONS with positions from the dummy template ----------
# Tab 2: header row = 12, detail rows 13..22
fill_tab2_region_hra_counts(
    template_path=TEMPLATE, output_path=OUTPUT, sheet_name="Tab 2",
    df_rows=df_tab2,
    region_col="region",
    metric_col="member_category",
    month_col="months",
    seq_col="seq_memb_id",
    header_row=12,
    first_detail_row=13, last_detail_row=22,
)

# Tab 3: no header; Jan starts at F (6); detail rows 90..115
fill_tab3_region_identity(
    template_path=TEMPLATE, output_path=OUTPUT, sheet_name="HRA Population Metrics",
    df_rows=df_tab3,
    region_col="region",
    category_col="category_type",
    subcat_col="sub_category",
    month_col="months",
    seq_col="seq_memb_id",
    start_col=6,
    first_detail_row=90, last_detail_row=115,
)

# Tab 4: header row = 30; detail rows 31..60
fill_tab4_identity_age(
    template_path=TEMPLATE, output_path=OUTPUT, sheet_name="Tab 4",
    df_rows=df_tab4,
    age_col="age",
    category_col="category_type",
    subcat_col="sub_category",
    month_col="months",
    seq_col="seq_memb_id",
    header_row=30,
    first_detail_row=31, last_detail_row=60,
)

# Tab 5: no header; Jan at D (4); detail rows 50..75
fill_tab5_region_needs(
    template_path=TEMPLATE, output_path=OUTPUT, sheet_name="Tab 5 - Region Needs",
    df_rows=df_tab5,
    region_col="region",
    need_col="HRA NEEDS ASSESSMENT METRICS",
    month_col="months",
    seq_col="seq_memb_id",
    start_col=4,
    first_detail_row=50, last_detail_row=75,
)

# Tab 6: super-simple totals — Jan at F(6), write to row 85
fill_tab6_month_totals_simple(
    template_path=TEMPLATE, output_path=OUTPUT, sheet_name="Tab 6 - Totals",
    df_rows=df_tab6,
    month_col="months", seq_col="seq_memb_id",
    jan_col=6, target_row=85,
    scope="quarter"  # or "year"
)

print("Dummy template filled ->", OUTPUT)
