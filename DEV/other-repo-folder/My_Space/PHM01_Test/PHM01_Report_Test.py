import os
import re
from contextlib import chdir

import configparser
#import MONTHS
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# from pandas._libs.tslibs.ccalendar import MONTHS
#
# (MONTHS)
# Create a configparser object

config = configparser.ConfigParser()
# os.chdir=("/Users/reinaldoburgos/sources/My_Space/my_projects/PHM01_Test/")
# os.listdir()

def report_var(parameters):
    config.read('PHM01_Test_New.ini')
    input_var = "file_input_path"
    output_var = "file_output_path"

    print(report_var)
    # os.getcwd(input_var)


#def main(input_var: str, file_one_path: None) -> object:
    # df = pd.read_csv("your_counts_source.csv")
    # Example skeleton:
    file_one_path = f'{input_var}file_one'
    df = pd.read_excel(file_one_path, sheet_name="Health Assessment(HRA)")
    print(file_one_path)

    return 0


# # ---------------- helpers ----------------
# def _load_wb(template_path: str, output_path: str):
#     """Open output if it exists; otherwise start from template."""
#     path = output_path if os.path.exists(output_path) else template_path
#     wb = load_workbook(path)
#     return wb
#
#
# def _save_wb(wb, output_path: str):
#     wb.save(output_path)
#
#
# def _normalize(s):
#     if s is None:
#         return ""
#     return str(s).strip()
#
#
# def _month_col_map_from_header(ws, header_row: int, start_col: int = 1, end_col: int = 200):
#     """Map month label -> column index by scanning a header row."""
#     m = {}
#     for c in range(start_col, end_col + 1):
#         v = _normalize(ws.cell(header_row, c).value)
#         if v:
#             m[v] = c
#     return m
#
#
# def _safe_write(ws, row: int, col: int, value):
#     """
#     If (row,col) is inside a merged range, write to the merged range's top-left.
#     Otherwise write to (row,col).
#     """
#     target_row, target_col = row, col
#     for rng in ws.merged_cells.ranges:
#         if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
#             target_row, target_col = rng.min_row, rng.min_col
#             break
#     ws.cell(target_row, target_col).value = value
#
#
# def _unique_count(df: pd.DataFrame, seq_col: str) -> int:
#     if df.empty:
#         return 0
#     return df[seq_col].nunique(dropna=True)
#
#
# def _parse_bucket(bucket: str):
#     """
#     Parse template bucket strings like '0-17', '18-24', '25-34', '65+'.
#     Returns (lo, hi) where hi can be None for '+'.
#     """
#     b = _normalize(bucket)
#     m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", b)
#     if m:
#         return int(m.group(1)), int(m.group(2))
#     m = re.match(r"^\s*(\d+)\s*\+\s*$", b)
#     if m:
#         return int(m.group(1)), None
#     return None
#
#
# def _age_to_bucket(age_value, available_buckets):
#     """
#     If age_value is already a bucket string like '25-34', return it if present.
#     If it's numeric, map it to one of the available buckets from the sheet.
#     """
#     # Already bucket-like
#     if isinstance(age_value, str):
#         v = _normalize(age_value)
#         if v in available_buckets:
#             return v
#         # if user gave '25 - 34' with spaces, normalize attempt
#         v2 = re.sub(r"\s+", "", v)
#         for b in available_buckets:
#             if re.sub(r"\s+", "", _normalize(b)) == v2:
#                 return b
#         return None
#
#     # Numeric
#     try:
#         a = int(age_value)
#     except Exception:
#         return None
#
#     # Build parsed buckets
#     parsed = []
#     for b in available_buckets:
#         p = _parse_bucket(b)
#         if p:
#             parsed.append((b, p[0], p[1]))
#
#     # Prefer exact range match
#     for b, lo, hi in parsed:
#         if hi is None:
#             if a >= lo:
#                 return b
#         else:
#             if lo <= a <= hi:
#                 return b
#     return None
#
#
# # ---------------- Tab 2 ----------------
#
# def fill_tab2_region_hra_counts(template_path, output_path, sheet_name, df_rows,
#                                 region_col, metric_col, month_col, seq_col,
#                                 header_row, first_detail_row, last_detail_row):
#     wb = _load_wb(template_path, output_path)
#     ws = wb[sheet_name]
#
#     # Month columns from header row (Jan..Dec usually starting col D)
#     month_to_col = _month_col_map_from_header(ws, header_row, start_col=1, end_col=200)
#
#     # Only keep months present in the sheet header
#     df = df_rows.copy()
#     df[month_col] = df[month_col].astype(str).str.strip()
#     df[region_col] = df[region_col].astype(str).str.strip()
#     df[metric_col] = df[metric_col].astype(str).str.strip()
#
#     for r in range(first_detail_row, last_detail_row + 1):
#         reg = _normalize(ws.cell(r, 1).value)
#         metric = _normalize(ws.cell(r, 2).value)
#
#         # skip blank/template separator rows
#         if not reg or not metric:
#             continue
#
#         for m, c in month_to_col.items():
#             if m not in MONTHS:
#                 continue
#             sub = df[(df[region_col] == reg) & (df[metric_col] == metric) & (df[month_col] == m)]
#             _safe_write(ws, r, c, _unique_count(sub, seq_col))
#
#     _save_wb(wb, output_path)
#
#
# # ---------------- Tab 3 ----------------
#
# def fill_tab3_region_identity(template_path, output_path, sheet_name, df_rows,
#                               region_col, category_col, subcat_col, month_col, seq_col,
#                               start_col, first_detail_row, last_detail_row):
#     wb = _load_wb(template_path, output_path)
#     ws = wb[sheet_name]
#
#     df = df_rows.copy()
#     df[month_col] = df[month_col].astype(str).str.strip()
#     df[region_col] = df[region_col].astype(str).str.strip()
#     df[category_col] = df[category_col].astype(str).str.strip()
#     df[subcat_col] = df[subcat_col].astype(str).str.strip()
#
#     for r in range(first_detail_row, last_detail_row + 1):
#         reg = _normalize(ws.cell(r, 1).value)
#         cat = _normalize(ws.cell(r, 2).value)
#         sub = _normalize(ws.cell(r, 3).value)
#
#         if not reg or not cat or not sub:
#             continue
#
#         # months horizontally from start_col
#         for i, m in enumerate(MONTHS):
#             c = start_col + i
#             subdf = df[(df[region_col] == reg) & (df[category_col].str.lower() == cat.lower()) &
#                        (df[subcat_col] == sub) & (df[month_col] == m)]
#             _safe_write(ws, r, c, _unique_count(subdf, seq_col))
#
#     _save_wb(wb, output_path)
#
#
# # ---------------- Tab 4 ----------------
#
# def fill_tab4_identity_age(template_path, output_path, sheet_name, df_rows,
#                            age_col, category_col, subcat_col, month_col, seq_col,
#                            header_row, first_detail_row, last_detail_row):
#     wb = _load_wb(template_path, output_path)
#     ws = wb[sheet_name]
#
#     month_to_col = _month_col_map_from_header(ws, header_row, start_col=1, end_col=200)
#
#     # Collect available age buckets from the sheet in the detail range
#     age_buckets = []
#     for r in range(first_detail_row, last_detail_row + 1):
#         v = _normalize(ws.cell(r, 1).value)
#         if v:
#             age_buckets.append(v)
#     age_buckets = sorted(set(age_buckets), key=lambda x: (len(x), x))
#
#     df = df_rows.copy()
#     df[month_col] = df[month_col].astype(str).str.strip()
#     df[category_col] = df[category_col].astype(str).str.strip()
#     df[subcat_col] = df[subcat_col].astype(str).str.strip()
#
#     # Map df age -> template bucket
#     df["_age_bucket_"] = df[age_col].apply(lambda x: _age_to_bucket(x, age_buckets))
#     df = df[df["_age_bucket_"].notna()].copy()
#
#     for r in range(first_detail_row, last_detail_row + 1):
#         age_bucket = _normalize(ws.cell(r, 1).value)
#         cat = _normalize(ws.cell(r, 2).value)
#         sub = _normalize(ws.cell(r, 3).value)
#
#         if not age_bucket or not cat or not sub:
#             continue
#
#         for m, c in month_to_col.items():
#             if m not in MONTHS:
#                 continue
#             subdf = df[(df["_age_bucket_"] == age_bucket) &
#                        (df[category_col].str.lower() == cat.lower()) &
#                        (df[subcat_col] == sub) &
#                        (df[month_col] == m)]
#             _safe_write(ws, r, c, _unique_count(subdf, seq_col))
#
#     _save_wb(wb, output_path)
#
#
# # ---------------- Tab 5 ----------------
#
# def fill_tab5_region_needs(template_path, output_path, sheet_name, df_rows,
#                            region_col, need_col, month_col, seq_col,
#                            start_col, first_detail_row, last_detail_row):
#     wb = _load_wb(template_path, output_path)
#     ws = wb[sheet_name]
#
#     df = df_rows.copy()
#     df[month_col] = df[month_col].astype(str).str.strip()
#     df[region_col] = df[region_col].astype(str).str.strip()
#     df[need_col] = df[need_col].astype(str).str.strip()
#
#     for r in range(first_detail_row, last_detail_row + 1):
#         reg = _normalize(ws.cell(r, 1).value)
#         need = _normalize(ws.cell(r, 2).value)
#
#         if not reg or not need:
#             continue
#
#         for i, m in enumerate(MONTHS):
#             c = start_col + i
#             subdf = df[(df[region_col] == reg) & (df[need_col] == need) & (df[month_col] == m)]
#             _safe_write(ws, r, c, _unique_count(subdf, seq_col))
#
#     _save_wb(wb, output_path)
#
#
# # ---------------- Tab 6 ----------------
#
# def fill_tab6_month_totals_simple(template_path, output_path, sheet_name, df_rows,
#                                   month_col, seq_col, jan_col, target_row, scope):
#     wb = _load_wb(template_path, output_path)
#     ws = wb[sheet_name]
#
#     df = df_rows.copy()
#     df[month_col] = df[month_col].astype(str).str.strip()
#
#     if str(scope).lower() == "quarter":
#         months = MONTHS[:3]  # Jan/Feb/Mar
#     else:
#         months = MONTHS[:]  # all year
#
#     for i, m in enumerate(months):
#         c = jan_col + i
#         subdf = df[df[month_col] == m]
#         _safe_write(ws, target_row, c, _unique_count(subdf, seq_col))
#
#     _save_wb(wb, output_path)
#

#***************************************************
#***************************************************
# main(
#     srcs=[0]
# )